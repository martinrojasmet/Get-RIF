from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
import tkinter as tk
from PIL import Image,ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from functions import check_format_rif, run_bot, set_values_xlsx, send_notification
from playwright.sync_api import Playwright, sync_playwright, expect

class GUI:
    def __init__(self, root):
        background_color = "#1E1E1E"
        text_color = "#AAAAAA"
        title_text_color = "#FFFFFF"
        run_button_color = "#40B85A"
        alert_text_color = "#C24A4A"
        excel_cells_color = "3268A8"
        excel_headers = [
            ("A", "RIF"),
            ("B", "CÉDULA O PASAPORTE"),
            ("C", "NOMBRE"),
            ("D", "DETALLES"),
            ("E", "% RETENCIÓN"),
            ("F", "MENSAJE"),
            ("G", "INTENTOS")
        ]

        self.xlsx_file = ""

        def adjust_upload_label_position(event):
            button_width = self.upload_button.winfo_width()
            button_height = self.upload_button.winfo_height()
            
            self.upload_label.place(
                in_=self.upload_button,
                x=button_width // 2,
                y=(button_height // 2) + 55,
                anchor='center'
            )
        
        def adjust_progress_bar_position(event):
            button_width = self.run_button.winfo_width()
            button_height = self.run_button.winfo_height()
            
            self.progress_bar.place(
                in_=self.run_button,
                x=button_width // 2,
                y=(button_height // 2) + 25,
                anchor='center',
                height=15
            )

        def adjust_info_position(event):
            width = self.root.winfo_width()
            height = self.root.winfo_height()

            
            self.info_button.place(
                x=width - 25,
                y=26,
                anchor='center'
            )

        def create_xlsx_file(folder_path):
            wb = Workbook()
            ws = wb.active

            fill = PatternFill(start_color=excel_cells_color, end_color=excel_cells_color, fill_type="solid")
            font = Font(name="Arial", bold=True, color="FFFFFF")
            alignment = Alignment(horizontal="center", vertical="center")

            for col, text in excel_headers:
                cell = ws[f"{col}1"]
                cell.value = text
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment
                ws.column_dimensions[col].width = len(text) + 5

            file_path = folder_path + "/" + "get_rif.xlsx"
            wb.save(file_path)
        
        def check_headers_file(file_path):
            result = True
            wb = load_workbook(file_path)
            ws = wb.active

            for col, text in excel_headers:
                cell = ws[f"{col}1"]
                if cell.value != text:
                    result = False
                    break
            
            wb.close()
            return result

        def file_not_empty(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            
            rif_cell = ws["A2"]
            cedula_cell = ws["B2"]

            wb.close()
            
            return rif_cell.value != None or cedula_cell.value != None

        def upload_files():
            try:
                file_path = filedialog.askopenfilename(initialdir = "./",
                                                    title = "Select a File",
                                                    filetypes = (("xlsx files",
                                                                    "*.xlsx*"),
                                                                ("xlsx files",
                                                                    "*.xlsx*")))
                if file_path:
                    file_type = file_path.split(".")[-1]
                    if file_type == "xlsx":
                        if file_not_empty(file_path) and check_headers_file(file_path):
                            self.upload_label.config(text="xlsx file selected", fg=text_color)
                            self.xlsx_file = file_path
                        else:
                            self.upload_label.config(text="Headers don't exists or file is empty", fg=alert_text_color)
                    else:
                        self.upload_label.config(text="Not a xlsx file", fg=alert_text_color)
                else:
                    self.upload_label.config(text="Error selecting the xlsx file", fg=alert_text_color)
            except Exception as e:
                print(e)
                self.upload_label.config(text="Error selecting the xlsx file", fg=alert_text_color)

        def select_folder():
            try:
                file_path = filedialog.askdirectory(initialdir = "./")
                create_xlsx_file(file_path)
                self.upload_label.config(text="xlsx file created", fg=text_color)
            except Exception as e:
                print(e)
                self.upload_label.config(text="Error creating xlsx file", fg=alert_text_color)

        def show_legend():
            # Create new window
            legend_window = tk.Toplevel(self.root)
            legend_window.title("Icon Legend")
            legend_window.configure(bg=background_color)
            
            # Create container frame
            legend_frame = tk.Frame(legend_window, bg=background_color)
            legend_frame.pack(padx=20, pady=20)

            # List of icon/text pairs
            legend_items = [
                ("./utils/icons/document_icon.png", """It creates an .xlsx file with the format
to run the program"""),
                ("./utils/icons/upload_icon.png", """You can upload the file that you want 
to process with the program"""),
                ("./utils/icons/info_icon.png", """The status is in the text below the upload
button, but you can visualize it in the
progress bar""")
            ]

            legend_window.legend_images = []  # 1. Attach to window object

            for row, (icon_path, text) in enumerate(legend_items):
                try: 
                    img = Image.open(icon_path)
                    img = img.resize((32, 32), Image.LANCZOS)
                    photo_img = ImageTk.PhotoImage(img)
                    legend_window.legend_images.append(photo_img)

                    icon_label = tk.Label(legend_frame, image=photo_img, bg=background_color)
                    icon_label.grid(row=row, column=0, padx=5, pady=5, sticky="w")

                    text_label = tk.Label(legend_frame, text=text, bg=background_color, 
                                        fg=text_color, font=("Arial", 10), justify=LEFT)
                    text_label.grid(row=row, column=1, padx=5, pady=5, sticky="w")
                    
                except Exception as e:
                    print(f"Error loading {icon_path}: {str(e)}")

        def calculate_progress(index, length):
            value = int((index + 1 / length) * 100)
            if value > 100:
                value = 100
            return value

        def run():
            try:
                if self.xlsx_file != "":
                    self.upload_label.config(text=f"Running on {self.xlsx_file}", fg=text_color)
                    try:
                        # send_notification("Get RIF", "El bot inició su ejecución")

                        df = pd.read_excel(self.xlsx_file)
                        df_length = len(df.index)
                        self.progress_bar["value"] = 0
                        self.progress_bar.update_idletasks()
                        print("Excel read")

                        for index, row in df.iterrows():
                            print("------------")
                            run_bot_bool = True
                            rif = row['RIF']
                            try:
                                cedula = str(int(row['CÉDULA O PASAPORTE']))
                            except Exception as e:
                                cedula = None

                            rif_exists = pd.notna(rif)
                            cedula_exists = pd.notna(cedula)

                            name = ""
                            details = ""
                            retention_percentage = ""
                            message = ""

                            if not rif_exists:
                                print("RIF does not exists")
                                run_bot_bool = False
                            else:
                                if not check_format_rif(rif):
                                    print(f"The RIF {rif} does not have the correct format")
                                    run_bot_bool = False
                            if not cedula_exists:
                                cedula = ""

                            if run_bot_bool:
                                with sync_playwright() as playwright:
                                    name, details, retention_percentage, attempts, message = run_bot(playwright, rif, cedula)
                                set_values_xlsx(row.name+2, name, details, retention_percentage, message, attempts)

                            self.progress_bar["value"] = calculate_progress(index, df_length)
                            self.progress_bar.update_idletasks()
                            
                        self.upload_label.config(text=f"Finished run on {self.xlsx_file}", fg=text_color)
                        # send_notification("Get RIF", "El bot terminó su ejecución de manera correcta", "check")
                    except Exception as e:
                        print(e)
                else:
                    self.upload_label.config(text="Error running", fg=alert_text_color)       
            except:
                self.upload_label.config(text="Error running", fg=alert_text_color)

        self.root = root
        self.root.geometry("600x350")
        self.root.config(bg=background_color)

        self.document_icon = ImageTk.PhotoImage(file = "./utils/icons/document_icon.png")
        self.document_button = tk.Button(self.root, image = self.document_icon, bg=background_color, border=0, 
                                        borderwidth=0, highlightthickness=0, activeforeground=background_color,
                                        activebackground=background_color, command=select_folder)
        self.document_button.pack()
        self.document_button.place(x=25, y=26, anchor=CENTER)

        img = Image.open("./utils/icons/info_icon.png")
        img = img.resize((30, 30), Image.LANCZOS)  # Adjust size as needed (width, height)
        self.info_icon = ImageTk.PhotoImage(img)
        self.info_button = tk.Button(self.root, image=self.info_icon, bg=background_color, border=0, 
                                        borderwidth=0, highlightthickness=0, activeforeground=background_color,
                                        activebackground=background_color, command=show_legend)
        self.info_button.pack()
        self.info_button.bind('<Configure>', adjust_info_position)

        self.get_rif_title = tk.Label(text="gRIF", fg=title_text_color, font=("Arial", 30))
        self.get_rif_title.configure(bg=background_color)
        self.get_rif_title.place(relx=0.5, y=40, anchor=CENTER)

        self.upload_icon = ImageTk.PhotoImage(file = "./utils/icons/upload_icon.png") 
        self.upload_button = tk.Button(self.root, image = self.upload_icon, bg=background_color, border=0, 
                                        borderwidth=0, highlightthickness=0, activeforeground=background_color,
                                        activebackground=background_color, command=upload_files)
        self.upload_button.pack()
        self.upload_button.place(relx=0.5, rely=0.5, anchor=CENTER)

        self.upload_label = tk.Label(text="Upload xlsx file...", fg=text_color)
        self.upload_label.configure(bg=background_color)
        self.upload_label.pack()
        self.upload_label.bind('<Configure>', adjust_upload_label_position)

        self.run_button = tk.Button(text="Run", bg=run_button_color, border=0, borderwidth=0, highlightthickness=0,
                                    width=8, command=run)
        self.run_button.pack()
        self.run_button.place(relx=0.5, rely=0.8, anchor=CENTER)

        self.progress_bar = Progressbar(length=200)
        self.progress_bar.pack()
        self.progress_bar.bind('<Configure>', adjust_progress_bar_position)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("gRIF")
    ico = Image.open('./utils/icons/gRIF.png')
    photo = ImageTk.PhotoImage(ico)
    root.wm_iconphoto(False, photo)
    gui = GUI(root)
    root.mainloop()