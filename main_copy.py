import pandas as pd
import re
from playwright.sync_api import Playwright, sync_playwright, expect
import easyocr
import requests
import cv2
import numpy as np
import re
import os
from openpyxl import load_workbook
from notifypy import Notify

base_url = "http://contribuyente.seniat.gob.ve/BuscaRif/"
file_path = 'get_rif.xlsx'

def check_format_rif(string):
    pattern = r'^[VJ][0-9]{9}$'
    result = False
    if re.match(pattern, string):
        result = True
    return result

def extract_name(text, rif):
    name = text.replace(rif, "").strip()
    return name

def edit_image_captcha():
    image = cv2.imread("tmp/captcha.jpg", 0)
    image = cv2.normalize(image, None, alpha=0, beta=255, norm_type=cv2.NORM_MINMAX)
    image = cv2.GaussianBlur(image,(1,3),0)
    umbral = 0.5 * 255
    ret, image = cv2.threshold(image, umbral, 255, cv2.THRESH_BINARY)
    image = cv2.dilate(image, np.ones((2, 2), np.uint8))
    image = cv2.bitwise_not(image)
    cv2.imwrite("tmp/captcha-fixed.jpg", image)

def add_detail(result, isFirst, text):
    if isFirst:
        result += text
        isFirst = False
    else:
        result += f"\n{text}"
    return result, isFirst

def define_details(text): 
    result = ""

    isFirst = True

    if("Contribuyente Ordinario del IVA" in text):
        # result += "-Es contribuyente ordinario del IVA"
        result, isFirst = add_detail(result, isFirst, "-Es contribuyente ordinario del IVA")
    if("Contribuyente Formal del IVA" in text):
        # result += "-Es contribuyente formal del IVA"
        result, isFirst = add_detail(result, isFirst, "-Es contribuyente formal del IVA")
    if("Agente de Retención del IVA" in text):
        # result += "-Es agente de retención del IVA"
        result, isFirst = add_detail(result, isFirst, "-Es agente de retención del IVA")

    return result

def run_bot(playwright: Playwright, rif, id) -> None:
    keepGoing = True
    attempts = 1

    # Go to page
    browser = playwright.chromium.launch(headless=True)

    context = browser.new_context()
    context.set_default_timeout(60000)
    cookies = context.cookies()

    page = context.new_page()
    page.goto(base_url + "BuscaRif.jsp")

    name = ""
    details = ""
    retention_percentage = ""
    message = ""

    while(keepGoing):
        # Fill Form
        page.get_by_role("textbox", name="Ingrese su número de Rif, seg").fill(rif)
        page.get_by_role("textbox", name="Ingrese su número de Cédula o").fill(id)

        # Download Image
        page.get_by_role("img").screenshot(path="tmp/captcha.jpg")

        # Edit Image
        edit_image_captcha()

        # Read Image
        reader = easyocr.Reader(['en'])
        result = reader.readtext('tmp/captcha-fixed.jpg', detail=0)

        # Screenshot of the captcha saved
        page.get_by_role("img").screenshot(path=f"utils/captcha_images/{result[0]}_NEW.jpg")

        # Fill Captcha
        page.locator("#codigo").fill(result[0])

        # Screenshot of the page
        page.screenshot(path="tmp/seniat.png")

        # Submit
        page.get_by_role("button", name="Buscar").click()

        errorExists = page.get_by_text("No existe el contribuyente").is_visible(timeout=2000)

        if errorExists:
            keepGoing = False
        else:
            keepGoing = page.get_by_text("EL código no coincide con la").is_visible()

        if (keepGoing):
            # Print error message
            print(f"Error at attempt number {attempts}")
            attempts += 1
        else:
            if errorExists:
                message = "No existe el contribuyente"
            else:
                # Save screenshot of the final page
                page.screenshot(path="tmp/seniat_final.png")

                # Get the text of the page
                name_text = page.get_by_text(rif).inner_text()
                name = extract_name(name_text, rif)
                text = page.get_by_text("Actividad Económica:").inner_text()

                # Get the percentage of retention
                retention_percentage = re.search(r'retención del (\d+)%', text)

                details = define_details(text)
                retention_percentage = f"{retention_percentage.group(1)}%"

    context.close()
    browser.close()

    return name, details, retention_percentage, attempts, message

def set_values_xlsx(row, name, details, retention_percentage, message, attempts):
    wb = load_workbook(file_path)
    ws = wb.active

    rif_col = 1
    cedula_col = 2
    nombre_col = 3
    tipo_col = 4
    retencion_col = 5
    mensaje_col = 6
    intentos_col = 7

    ws.cell(row=row, column=nombre_col, value=name)
    ws.cell(row=row, column=tipo_col, value=details)
    ws.cell(row=row, column=retencion_col, value=retention_percentage)
    ws.cell(row=row, column=mensaje_col, value="Ejecución sin errores" if message == "" else message)
    ws.cell(row=row, column=intentos_col, value=attempts)

    wb.save(file_path)

def send_notification(title, message, icon_name=None):
    notification = Notify()
    notification.title = title
    notification.message = message
    if icon_name:
        notification.icon = f"./utils/icons/{icon_name}.png"

    notification.send()

def run():
    send_notification("Get RIF", "El bot inició su ejecución")

    df = pd.read_excel(file_path)
    print("Excel read")

    for index, row in df.iterrows():
        print("------------")
        rif = row['RIF']
        print(rif)
        try:
            cedula = str(int(row['CÉDULA O PASAPORTE']))
        except Exception as e:
            cedula = None

        rif_exists = pd.notna(rif)
        cedula_exists = pd.notna(cedula)

        name = ""
        details = ""
        retention_percentage = ""
        message = ""

        if not rif_exists:
            print("RIF does not exists")
            continue
        else:
            if not check_format_rif(rif):
                print(f"The RIF {rif} does not have the correct format")
                continue
        if not cedula_exists:
            cedula = ""

        with sync_playwright() as playwright:
            name, details, retention_percentage, attempts, message = run_bot(playwright, rif, cedula)

        set_values_xlsx(row.name+2, name, details, retention_percentage, message, attempts)
        print("Values set in the excel file")

    send_notification("Get RIF", "El bot terminó su ejecución de manera correcta", "check")

try:
    run()
except:
    send_notification("Get RIF", "Ocurrió un error", "check")
